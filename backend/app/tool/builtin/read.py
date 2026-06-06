"""Read tool — read file contents with optional offset/limit paging."""

from __future__ import annotations

import csv
import io
import logging
import os
from pathlib import Path
from typing import Any

from app.tool.base import ToolDefinition, ToolResult
from app.tool.context import ToolContext
from app.tool.extractors import extract_document, is_supported_binary
from app.tool.workspace import WorkspaceViolation, resolve_and_validate

logger = logging.getLogger(__name__)

_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".webp", ".svg"}
_DATA_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}
_DATA_SAMPLE_ROWS = 5


class ReadTool(ToolDefinition):

    @property
    def id(self) -> str:
        return "read"

    @property
    def is_concurrency_safe(self) -> bool:
        return True

    @property
    def description(self) -> str:
        return (
            "Read a file from the filesystem. Supports offset/limit for paging "
            "through large files. Can also list directory contents. "
            "Natively handles ALL file types including PDF, DOCX, XLSX, PPTX, "
            "and images (PNG, JPG, etc.) — no skill or plugin needed, just call read directly. "
            "Use 'pages' to read specific pages of a PDF or a named sheet in XLSX. "
            "Use 'format=json' for structured spreadsheet output."
        )

    def parameters_schema(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Absolute or relative path to the file to read",
                },
                "offset": {
                    "type": "integer",
                    "description": "Line number to start reading from (1-based)",
                    "default": 1,
                },
                "limit": {
                    "type": "integer",
                    "description": "Maximum number of lines to read",
                    "default": 2000,
                },
                "pages": {
                    "type": "string",
                    "description": "Page range for PDFs/PPTX (e.g. '1-3' or '5'), or sheet name for XLSX (e.g. 'Revenue')",
                },
                "format": {
                    "type": "string",
                    "enum": ["json"],
                    "description": "Set to 'json' for structured spreadsheet output (XLSX/CSV)",
                },
            },
            "required": ["file_path"],
        }

    async def execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        file_path = args.get("file_path", "")
        ext = os.path.splitext(file_path)[1].lower()

        # Images: return as base64 for multimodal LLM
        if ext in _IMAGE_EXTENSIONS:
            return self._read_image(file_path, ctx)

        # Data files: return schema + sample only
        if ext in _DATA_EXTENSIONS:
            return self._read_data_file(file_path, ext, ctx)

        return await self._filesystem_execute(args, ctx)

    # ------------------------------------------------------------------
    # Data files: schema + sample (CSV, XLSX)
    # ------------------------------------------------------------------

    def _read_data_file(self, file_path: str, ext: str, ctx: ToolContext) -> ToolResult:
        """Return schema and a few sample rows for data files.

        For actual analysis the agent should use code_execute with pandas.
        """
        try:
            resolved = resolve_and_validate(file_path, ctx.workspace)
        except WorkspaceViolation as e:
            return ToolResult(error=str(e))

        if not os.path.exists(resolved):
            return ToolResult(error=f"File not found: {file_path}")

        try:
            if ext in (".csv", ".tsv"):
                return self._summarise_csv(resolved, file_path)
            else:
                return self._summarise_xlsx(resolved, file_path)
        except Exception as e:
            return ToolResult(error=f"Cannot read {os.path.basename(file_path)}: {e}")

    def _summarise_csv(self, resolved: str, file_path: str) -> ToolResult:
        with open(resolved, "r", encoding="utf-8", errors="replace") as f:
            # Sniff delimiter
            sample_text = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample_text)
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = "," if file_path.lower().endswith(".csv") else "\t"

            reader = csv.reader(f, delimiter=delimiter)
            rows: list[list[str]] = []
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= _DATA_SAMPLE_ROWS:  # header + N sample rows
                    break

        if not rows:
            return ToolResult(output="(Empty file)", title=os.path.basename(file_path))

        headers = rows[0]
        data_rows = rows[1:]

        # Count total rows without reading everything into memory
        with open(resolved, "r", encoding="utf-8", errors="replace") as f:
            total_rows = sum(1 for _ in f) - 1  # minus header

        # Build output
        parts = [
            f"File: {os.path.basename(file_path)}",
            f"Rows: {total_rows:,}  |  Columns: {len(headers)}",
            "",
            "Columns: " + ", ".join(headers),
            "",
            "Sample rows:",
        ]
        # Header
        parts.append(" | ".join(headers))
        parts.append(" | ".join(["---"] * len(headers)))
        for row in data_rows:
            # Pad/truncate to match header count
            padded = row[:len(headers)] + [""] * max(0, len(headers) - len(row))
            parts.append(" | ".join(padded))

        parts.append("")
        parts.append("Use code_execute for analysis.")

        return ToolResult(
            output="\n".join(parts),
            title=os.path.basename(file_path),
            metadata={
                "source": "filesystem",
                "format": "csv",
                "total_rows": total_rows,
                "columns": headers,
            },
        )

    def _summarise_xlsx(self, resolved: str, file_path: str) -> ToolResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is not installed")

        wb = load_workbook(resolved, read_only=True, data_only=True)
        parts = [f"File: {os.path.basename(file_path)}", f"Sheets: {', '.join(wb.sheetnames)}", ""]

        all_metadata: dict[str, Any] = {"source": "filesystem", "format": "xlsx", "sheets": {}}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(cells)
                if i >= _DATA_SAMPLE_ROWS:
                    break

            if not rows:
                parts.append(f"=== {sheet_name}: (empty) ===")
                continue

            headers = rows[0]
            data_rows = rows[1:]
            total_rows = ws.max_row - 1 if ws.max_row else 0

            parts.append(f"=== {sheet_name} ({total_rows:,} rows, {len(headers)} cols) ===")
            parts.append("Columns: " + ", ".join(headers))
            parts.append("")
            parts.append(" | ".join(headers))
            parts.append(" | ".join(["---"] * len(headers)))
            for row in data_rows:
                padded = row[:len(headers)] + [""] * max(0, len(headers) - len(row))
                parts.append(" | ".join(padded))
            parts.append("")

            all_metadata["sheets"][sheet_name] = {
                "total_rows": total_rows,
                "columns": headers,
            }

        wb.close()

        parts.append("Use code_execute for analysis.")

        return ToolResult(
            output="\n".join(parts),
            title=os.path.basename(file_path),
            metadata=all_metadata,
        )

    # ------------------------------------------------------------------
    # Image path (base64 for multimodal LLM)
    # ------------------------------------------------------------------

    def _read_image(self, file_path: str, ctx: ToolContext) -> ToolResult:
        """Return an image as base64 for the LLM to see visually.

        Stores the data URL in metadata so the message builder can convert
        the tool result into multimodal content (text + image_url).
        """
        import base64
        import mimetypes

        # Resolve path — try workspace-relative first, then absolute
        try:
            resolved = resolve_and_validate(file_path, ctx.workspace)
        except WorkspaceViolation:
            # Attachments may live outside the workspace — allow if file exists
            resolved = str(Path(file_path).resolve())

        if not os.path.exists(resolved):
            return ToolResult(error=f"Image not found: {file_path}")

        try:
            raw = Path(resolved).read_bytes()
            b64 = base64.b64encode(raw).decode("utf-8")
            mime_type, _ = mimetypes.guess_type(resolved)
            if not mime_type or not mime_type.startswith("image/"):
                ext = os.path.splitext(resolved)[1].lstrip(".")
                mime_type = f"image/{ext}"

            data_url = f"data:{mime_type};base64,{b64}"

            return ToolResult(
                output=f"[Image: {os.path.basename(file_path)}]",
                title=os.path.basename(file_path),
                metadata={
                    "source": "filesystem",
                    "format": os.path.splitext(file_path)[1].lower(),
                    "image_data_url": data_url,
                },
            )
        except Exception as e:
            return ToolResult(error=f"Cannot read image {file_path}: {e}")

    # ------------------------------------------------------------------
    # Filesystem fallback path
    # ------------------------------------------------------------------

    async def _filesystem_execute(self, args: dict[str, Any], ctx: ToolContext) -> ToolResult:
        file_path = args["file_path"]

        # Workspace restriction check
        try:
            file_path = resolve_and_validate(file_path, ctx.workspace)
        except WorkspaceViolation as e:
            return ToolResult(error=str(e))

        offset = max(1, args.get("offset", 1))
        limit = args.get("limit", 2000)

        if not os.path.exists(file_path):
            return ToolResult(error=f"File not found: {file_path}")

        # Directory listing
        if os.path.isdir(file_path):
            try:
                entries = sorted(os.listdir(file_path))
                listing = "\n".join(entries)
                return ToolResult(
                    output=listing,
                    title=f"Listed {len(entries)} entries in {os.path.basename(file_path)}",
                    metadata={"source": "filesystem"},
                )
            except PermissionError:
                return ToolResult(error=f"Permission denied: {file_path}")

        # Binary document extraction (PDF, DOCX, XLSX, PPTX)
        if is_supported_binary(file_path):
            try:
                text = extract_document(file_path)
            except ImportError as e:
                return ToolResult(error=str(e))
            except Exception as e:
                return ToolResult(
                    error=f"Cannot read {os.path.basename(file_path)}: {e}"
                )

            result = self._format_lines(text, file_path, offset, limit)
            if result.metadata is None:
                result.metadata = {}
            result.metadata["source"] = "filesystem"
            return result

        # Text file reading
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                lines = f.readlines()

            total_lines = len(lines)

            # Apply offset (1-based) and limit
            start = offset - 1
            end = start + limit
            selected = lines[start:end]

            # Format with line numbers (cat -n style)
            output_lines = []
            for i, line in enumerate(selected, start=offset):
                # Truncate very long lines
                line_content = line.rstrip("\n\r")
                if len(line_content) > 2000:
                    line_content = line_content[:2000] + "..."
                output_lines.append(f"{i:>6}\t{line_content}")

            output = "\n".join(output_lines)

            if end < total_lines:
                output += f"\n\n... ({total_lines - end} more lines)"

            return ToolResult(
                output=output,
                title=os.path.basename(file_path),
                metadata={"total_lines": total_lines, "shown": len(selected), "source": "filesystem"},
            )

        except UnicodeDecodeError:
            return ToolResult(error=f"Cannot read binary file: {file_path}")
        except PermissionError:
            return ToolResult(error=f"Permission denied: {file_path}")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _format_lines(
        text: str, file_path: str, offset: int, limit: int
    ) -> ToolResult:
        """Format extracted text with line numbers, applying offset/limit."""
        lines = text.split("\n")
        total_lines = len(lines)

        start = offset - 1
        end = start + limit
        selected = lines[start:end]

        output_lines = []
        for i, line in enumerate(selected, start=offset):
            line_content = line.rstrip("\n\r")
            if len(line_content) > 2000:
                line_content = line_content[:2000] + "..."
            output_lines.append(f"{i:>6}\t{line_content}")

        output = "\n".join(output_lines)

        if end < total_lines:
            output += f"\n\n... ({total_lines - end} more lines)"

        ext = os.path.splitext(file_path)[1].lower()
        return ToolResult(
            output=output,
            title=os.path.basename(file_path),
            metadata={
                "total_lines": total_lines,
                "shown": len(selected),
                "format": ext,
            },
        )
